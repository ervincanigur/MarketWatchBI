import openpyxl
from openpyxl.styles import Font, Alignment


class Populate(object):

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def write(self, data):
        for row, val in enumerate(data, start=1):
            for col, v in enumerate(val, start=1):
                self.ws.cell(row, col, v)
                if row == 1:  # Bolds first row
                    self.ws.cell(row, col).font = Font(b=True)
                elif col == 2:  # Center the dates'
                    self.ws.cell(row, col).alignment = Alignment(horizontal='center')
        self.adjust_cell_width()

    def adjust_cell_width(self):
        # Width adjustment
        for col in self.ws.columns:
            column_name = col[0].column_letter
            max_length = len(col[0].value)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            new_width = max_length * 1.2
            self.ws.column_dimensions[column_name].width = new_width

    def save(self, file):
        self.wb.save(file)
